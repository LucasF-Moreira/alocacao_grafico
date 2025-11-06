import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# -----------------------------
# CONFIGURAÇÕES INICIAIS
# -----------------------------
st.set_page_config(page_title="Gantt de Alocação", layout="wide")
st.title("📊 Painel de Alocação de Pessoas")
st.write("Visualize as etapas dos processos com alocação diária e filtros personalizados.")

# -----------------------------
# UPLOAD
# -----------------------------
arquivo = st.file_uploader("Envie a base Excel (.xlsx)", type=["xlsx"])
if arquivo:
    df = pd.read_excel(arquivo)
    df.columns = [c.strip() for c in df.columns]

    # -----------------------------
    # FILTROS GERAIS
    # -----------------------------
    col1, col2, col3, col4 = st.columns(4)
    escala = col1.selectbox("Escala de tempo", ["Diário", "Quinzenal"])
    tipo_pessoa = col2.selectbox("Tipo de pessoa", ["Todos", "Head", "Analista"])
    data_inicio = pd.to_datetime(col3.date_input("Data início"))
    data_fim = pd.to_datetime(col4.date_input("Data fim"))

    # -----------------------------
    # PREPARAÇÃO DOS DADOS
    # -----------------------------
    etapas = [1, 2, 3]
    registros = []
    for _, row in df.iterrows():
        for etapa in etapas:
            ini = row.get(f"Data_Inicio_Etapa{etapa}")
            fim = row.get(f"Data_Termino_Etapa{etapa}")
            if pd.notna(ini) and pd.notna(fim):
                consultor = row.get(f"Consultor_Head_Etapa_{etapa}_PCT", 0)
                analista = row.get(f"Analista_Etapa_{etapa}_PCT", 0)

                if pd.notna(row.get("Head_1")):
                    registros.append({
                        "Pessoa": row["Head_1"],
                        "Processo": row["Nome_Programa"],
                        "Etapa": f"Etapa {etapa}",
                        "Data_Inicio": ini,
                        "Data_Fim": fim,
                        "Pct": consultor,
                        "Tipo": "Head"
                    })
                if pd.notna(row.get("Analista_1")):
                    registros.append({
                        "Pessoa": row["Analista_1"],
                        "Processo": row["Nome_Programa"],
                        "Etapa": f"Etapa {etapa}",
                        "Data_Inicio": ini,
                        "Data_Fim": fim,
                        "Pct": analista,
                        "Tipo": "Analista"
                    })

    df_melt = pd.DataFrame(registros)
    if df_melt.empty:
        st.warning("Nenhum dado encontrado após processar as etapas.")
        st.stop()

    # -----------------------------
    # APLICA FILTROS
    # -----------------------------
    df_melt["Data"] = pd.to_datetime(df_melt["Data_Inicio"])
    df_melt["Data_Fim"] = pd.to_datetime(df_melt["Data_Fim"])
    df_melt = df_melt[(df_melt["Data_Fim"] >= data_inicio) & (df_melt["Data"] <= data_fim)]

    if tipo_pessoa == "Head":
        df_melt = df_melt[df_melt["Tipo"] == "Head"]
    elif tipo_pessoa == "Analista":
        df_melt = df_melt[df_melt["Tipo"] == "Analista"]

    if df_melt.empty:
        st.warning("Nenhum dado disponível após aplicar os filtros selecionados.")
        st.stop()

    # -----------------------------
    # GERA DATAS DIÁRIAS
    # -----------------------------
    all_rows = []
    for _, row in df_melt.iterrows():
        for d in pd.date_range(row["Data"], row["Data_Fim"]):
            if data_inicio <= d <= data_fim:
                all_rows.append({
                    "Pessoa": row["Pessoa"],
                    "Processo": row["Processo"],
                    "Etapa": row["Etapa"],
                    "Data": d,
                    "Pct": row["Pct"],
                    "Tipo": row["Tipo"]
                })
    df_f = pd.DataFrame(all_rows)

    # -----------------------------
    # ESCALA + LABELS ORDENADAS
    # -----------------------------
    df_f["Data"] = pd.to_datetime(df_f["Data"], errors="coerce")
    df_f["Dia_Real"] = df_f["Data"]

    if escala == "Diário":
        df_f["Data_label"] = df_f["Data"].dt.strftime("%d/%b")
    else:
        def quinzena_label(d):
            return f"{d.strftime('%b/%Y')} - {'1ª' if d.day <= 15 else '2ª'}"
        df_f["Data_label"] = df_f["Data"].apply(quinzena_label)

    # -----------------------------
    # CORES
    # -----------------------------
    def cor_por_etapa(etapa):
        if "1" in str(etapa):
            return "#90CAF9"
        elif "2" in str(etapa):
            return "#A5D6A7"
        elif "3" in str(etapa):
            return "#FFF59D"
        else:
            return "#E0E0E0"

    df_f["Cor"] = df_f["Etapa"].apply(cor_por_etapa)
    # agora mantemos o percentual completo, multiplicado por 100
    df_f["CorPct"] = df_f.apply(
        lambda x: f"{x['Cor']}|{x['Pct']*100:.2f}" if pd.notna(x['Pct']) else f"{x['Cor']}|",
        axis=1
    )

    # -----------------------------
    # FILTRO DE PESSOA
    # -----------------------------
    pessoas_unicas = sorted(df_f["Pessoa"].unique())
    pessoas_sel = st.multiselect("Selecione pessoas", ["Selecionar todos"] + pessoas_unicas, default=["Selecionar todos"])
    if "Selecionar todos" not in pessoas_sel:
        df_f = df_f[df_f["Pessoa"].isin(pessoas_sel)]

    # -----------------------------
    # PIVOT ORDENADO
    # -----------------------------
    ordem_datas = sorted(df_f["Dia_Real"].dt.date.unique())
    ordem_labels = []
    for d in ordem_datas:
        if escala == "Diário":
            label = pd.to_datetime(d).strftime("%d/%b")
        else:
            label = f"{pd.to_datetime(d).strftime('%b/%Y')} - {'1ª' if pd.to_datetime(d).day <= 15 else '2ª'}"
        if label not in ordem_labels:
            ordem_labels.append(label)

    pivot = df_f.pivot_table(
        index=["Pessoa", "Processo", "Etapa"],
        columns="Data_label",
        values="CorPct",
        aggfunc="first",
        fill_value="#FFFFFF|"
    ).reset_index()

    ordered_cols = ["Pessoa", "Processo", "Etapa"] + [c for c in ordem_labels if c in pivot.columns]
    pivot = pivot[ordered_cols]

    # -----------------------------
    # ZOOM
    # -----------------------------
    zoom = st.slider("Zoom (largura das colunas de data)", 40, 200, 80, 5)

    # -----------------------------
    # GRID CONFIG
    # -----------------------------
    gb = GridOptionsBuilder.from_dataframe(pivot)
    gb.configure_grid_options(domLayout='normal', suppressRowTransform=True)

    custom_css = {
        ".ag-header-cell-label": {
            "justify-content": "center",
            "font-size": "10px",
            "white-space": "normal",
            "line-height": "1.1em"
        },
        ".ag-cell": {
            "font-size": "11px",
            "text-align": "center",
            "padding": "0px"
        }
    }

    cell_color = JsCode("""
        function(params) {
            if (!params.value) return {'backgroundColor': '#FFFFFF', 'color': '#000000'};
            var parts = params.value.split('|');
            var bg = parts[0];
            return {
                'backgroundColor': bg,
                'color': '#000000',
                'textAlign': 'center',
                'fontWeight': 'bold',
                'border': '0.5px solid #777'
            };
        }
    """)

    cell_renderer = JsCode("""
        function(params) {
            if (!params.value) return '';
            var parts = params.value.split('|');
            if (!parts[1]) return '';
            let pct = parseFloat(parts[1]);
            if (isNaN(pct)) return '';
            return pct.toFixed(2) + '%';
        }
    """)

    date_cols = [c for c in pivot.columns if c in ordem_labels]
    for col in date_cols:
        gb.configure_column(
            col,
            cellStyle=cell_color,
            cellRenderer=cell_renderer,
            minWidth=max(45, zoom - 15),
            maxWidth=zoom + 15,
            resizable=True,
            autoSize=True
        )

    gb.configure_column("Pessoa", pinned="left")
    gb.configure_column("Processo", pinned="left")
    gb.configure_column("Etapa", pinned="left")

    grid_options = gb.build()

    # -----------------------------
    # EXIBE GANTT
    # -----------------------------
    st.markdown("### 🗓️ Gantt de Alocação por Pessoa")
    AgGrid(
        pivot,
        gridOptions=grid_options,
        update_mode=GridUpdateMode.NO_UPDATE,
        theme="alpine",
        allow_unsafe_jscode=True,
        fit_columns_on_grid_load=False,
        autoSizeAllColumns=True,
        custom_css=custom_css,
        height=750,
        enable_enterprise_modules=True,
    )

    # -----------------------------
    # EXPORTAR EXCEL COLORIDO
    # -----------------------------
    export_df = pivot.copy()
    for c in export_df.columns:
        if any(x in c for x in ["/", "ª"]):
            export_df[c] = export_df[c].apply(lambda v: v.split("|")[1] if isinstance(v, str) and "|" in v else "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt"

    for col_idx, col_name in enumerate(export_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx in range(len(export_df)):
        for col_idx, col_name in enumerate(export_df.columns, 1):
            value = export_df.iloc[row_idx, col_idx - 1]
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

            if col_name in date_cols and value not in ["", None, "0", 0]:
                etapa = export_df.iloc[row_idx, 2]  # coluna Etapa
                cor = cor_por_etapa(etapa)
                cell.fill = PatternFill(start_color=cor.replace("#", ""), end_color=cor.replace("#", ""), fill_type="solid")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    st.download_button(
        label="💾 Baixar Gantt em Excel (colorido)",
        data=buffer,
        file_name="Gantt_Alocacao.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

